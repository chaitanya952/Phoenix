#!/usr/bin/env python3
"""
Script to extract all products from Phoenix Catalog 7 Excel file and update the website catalog
with all 10 categories while preserving the current structure
"""

import sys
import os
from openpyxl import load_workbook

def clean_value(value):
    """Clean and convert cell values"""
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value)

def convert_to_number(value, default=0):
    """Convert value to number, return default if conversion fails"""
    if value is None or value == "":
        return default
    try:
        # Handle string values
        if isinstance(value, str):
            value = value.strip()
            if value == "" or value.upper() == "NA":
                return default
        
        # Try to convert to float first, then int if it's a whole number
        num = float(value)
        if num.is_integer():
            return int(num)
        return num
    except (ValueError, TypeError):
        return default

def extract_products_from_sheet(sheet, sheet_name):
    """Extract products from a single sheet"""
    products = []
    
    # Get headers from first row
    headers = []
    for col in range(1, sheet.max_column + 1):
        header = clean_value(sheet.cell(row=1, column=col).value)
        headers.append(header)
    
    print(f"Processing {sheet_name} - {sheet.max_row - 1} potential products")
    
    # Process each data row
    for row in range(2, sheet.max_row + 1):
        row_data = {}
        has_data = False
        
        # Extract data for each column
        for col in range(1, len(headers) + 1):
            if col <= sheet.max_column:
                value = clean_value(sheet.cell(row=row, column=col).value)
                row_data[headers[col-1]] = value
                if value and value != "":
                    has_data = True
        
        # Skip empty rows or rows without model number
        if not has_data or not row_data.get('Model No.', '').strip():
            continue
        
        # Create product object
        product = {
            'modelNo': row_data.get('Model No.', ''),
            'description': row_data.get('Description', ''),
            'packaging': row_data.get('Packaging', ''),
            'inner': convert_to_number(row_data.get('Inner', 0)),
            'outer': convert_to_number(row_data.get('Outer', 0)),
            'inr': convert_to_number(row_data.get('INR', 0)),
            'usd': convert_to_number(row_data.get('USD', 0)),
            'moq': convert_to_number(row_data.get('MOQ', 3600))
        }
        
        # Add optional fields if they exist
        if 'S.No' in row_data and row_data['S.No']:
            product['sNo'] = convert_to_number(row_data['S.No'])
        
        if 'Capacity' in row_data and row_data['Capacity']:
            product['capacity'] = row_data['Capacity']
        
        if 'Mold Nos.' in row_data and row_data['Mold Nos.']:
            product['moldNos'] = row_data['Mold Nos.']
        
        if 'Remarks' in row_data and row_data['Remarks']:
            product['remarks'] = row_data['Remarks']
        
        # Add dimension fields - check for various dimension column names
        for key in row_data.keys():
            if key and row_data[key]:
                if 'mono' in key.lower() and 'dimension' in key.lower():
                    product['monoDimensions'] = row_data[key]
                elif 'inner' in key.lower() and 'dimension' in key.lower():
                    product['innerDimensions'] = row_data[key]
                elif 'master' in key.lower() and 'dimension' in key.lower():
                    product['masterDimensions'] = row_data[key]
                elif 'cbm' in key.lower():
                    product['cbm'] = convert_to_number(row_data[key])
        
        products.append(product)
    
    return products

def create_category_mapping():
    """Create mapping from sheet names to category info"""
    return {
        'Std. Feeding': {
            'id': 'std-feeding',
            'name': 'Std. Feeding',
            'displayName': 'Standard Neck Feeding Bottles',
            'description': 'Classic standard neck bottles with reliable performance and proven design. Perfect for everyday feeding with universal compatibility.',
            'image': '/images/Standard Neck Bottles JPEG/RN0001 - 125ml.jpg'
        },
        'Wide Feeding': {
            'id': 'wide-feeding',
            'name': 'Wide Feeding',
            'displayName': 'Wide Neck Feeding Bottles',
            'description': 'Premium wide neck feeding bottles with superior design and easy cleaning. Perfect for comfortable feeding experience.',
            'image': '/images/Wide Neck Bottles JPEG/WN0001 - 210ml.jpg'
        },
        'Sippy Cups': {
            'id': 'sippy-cups',
            'name': 'Sippy Cups',
            'displayName': 'Sippy Cups & Training Cups',
            'description': 'Transition cups for toddlers with spill-proof design and various capacity options. Perfect for developing independence.',
            'image': '/images/Sippy Cups JPEG/SP0001 - 150ml.jpg'
        },
        'Teethers & Pacifiers': {
            'id': 'teethers-pacifiers',
            'name': 'Teethers & Pacifiers',
            'displayName': 'Teethers & Pacifiers',
            'description': 'Safe and soothing teethers and pacifiers for baby comfort and development. Made with food-grade materials.',
            'image': '/images/Teethers & Pacifiers JPEG/TP0001.JPG'
        },
        'Cutlery & Tableware': {
            'id': 'cutlery-tableware',
            'name': 'Cutlery & Tableware',
            'displayName': 'Cutlery & Tableware',
            'description': 'Complete feeding solutions including spoons, forks, bowls and plates for solid food introduction.',
            'image': '/images/Cutlery & Tableware JPEG/CT0001.JPG'
        },
        'MotherCare': {
            'id': 'mothercare',
            'name': 'MotherCare',
            'displayName': 'MotherCare Products',
            'description': 'Essential products for nursing mothers including breast pumps, nipple shields and accessories.',
            'image': '/images/MotherCare JPEG/MC0001.JPG'
        },
        'Others': {
            'id': 'others',
            'name': 'Others',
            'displayName': 'Other Baby Products',
            'description': 'Additional baby care products including bottle accessories, cleaning tools and storage solutions.',
            'image': '/images/Others JPEG/OT0001.JPG'
        },
        'Gift Sets': {
            'id': 'gift-sets',
            'name': 'Gift Sets',
            'displayName': 'Gift Sets & Combos',
            'description': 'Thoughtfully curated gift sets and product combinations perfect for baby showers and new parents.',
            'image': '/images/Gift Sets JPEG/GT0001.JPG'
        },
        'Spare Parts': {
            'id': 'spare-parts',
            'name': 'Spare Parts',
            'displayName': 'Spare Parts & Accessories',
            'description': 'Replacement parts and accessories for feeding bottles including nipples, caps, and other components.',
            'image': '/images/Spare Parts JPEG/SP0001.JPG'
        },
        'Glow in the Dark Series': {
            'id': 'glow-in-dark',
            'name': 'Glow in the Dark Series',
            'displayName': 'Glow in the Dark Series',
            'description': 'Special glow-in-the-dark products including pacifiers and teethers that provide comfort during nighttime.',
            'image': '/images/Missing Images JPEG/GT0001.JPG'
        }
    }

def extract_all_products(file_path):
    """Extract all products from the Excel file"""
    try:
        workbook = load_workbook(file_path, data_only=True)
        category_mapping = create_category_mapping()
        
        all_categories = []
        
        print("=== EXTRACTING PRODUCTS FROM PHOENIX CATALOG 7 ===")
        
        # Process each relevant sheet
        for sheet_name in workbook.sheetnames:
            if sheet_name in category_mapping:
                sheet = workbook[sheet_name]
                products = extract_products_from_sheet(sheet, sheet_name)
                
                if products:
                    category_info = category_mapping[sheet_name]
                    category = {
                        'id': category_info['id'],
                        'name': category_info['name'],
                        'displayName': category_info['displayName'],
                        'description': category_info['description'],
                        'image': category_info['image'],
                        'products': products
                    }
                    all_categories.append(category)
                    print(f"✓ {sheet_name}: {len(products)} products extracted")
                else:
                    print(f"⚠ {sheet_name}: No products found")
        
        return all_categories
        
    except Exception as e:
        print(f"Error extracting products: {e}")
        return []

def generate_typescript_catalog(categories):
    """Generate TypeScript catalog file"""
    
    # Start of the TypeScript file
    ts_content = """// src/data/catalogData.ts
export interface CatalogProduct {
  sNo?: number;
  modelNo: string;
  image?: string;
  description: string;
  capacity?: string;
  moldNos?: string;
  remarks?: string;
  packaging: string;
  inner: number;
  outer: number;
  inr: number;
  usd: number;
  moq: number;
  dimensions?: string;
  cbm?: number;
  monoDimensions?: string;
  innerDimensions?: string;
  masterDimensions?: string;
}

export interface ProductCategory {
  id: string;
  name: string;
  displayName: string;
  description: string;
  image: string;
  products: CatalogProduct[];
}

export const PRODUCT_CATEGORIES: ProductCategory[] = [
"""
    
    # Add each category
    for i, category in enumerate(categories):
        ts_content += f"  {{\n"
        ts_content += f"    id: '{category['id']}',\n"
        ts_content += f"    name: '{category['name']}',\n"
        ts_content += f"    displayName: '{category['displayName']}',\n"
        ts_content += f"    description: '{category['description']}',\n"
        ts_content += f"    image: '{category['image']}',\n"
        ts_content += f"    products: [\n"
        
        # Add products
        for j, product in enumerate(category['products']):
            ts_content += f"      {{\n"
            
            # Add all product fields
            for key, value in product.items():
                if isinstance(value, str):
                    # Escape single quotes in strings
                    escaped_value = value.replace("'", "\\'")
                    ts_content += f"        {key}: '{escaped_value}',\n"
                else:
                    ts_content += f"        {key}: {value},\n"
            
            ts_content += f"      }}"
            if j < len(category['products']) - 1:
                ts_content += ","
            ts_content += "\n"
        
        ts_content += f"    ]\n"
        ts_content += f"  }}"
        if i < len(categories) - 1:
            ts_content += ","
        ts_content += "\n"
    
    # End of the TypeScript file
    ts_content += """];

// Helper function to get category by ID
export const getCategoryById = (id: string): ProductCategory | undefined => {
  return PRODUCT_CATEGORIES.find(category => category.id === id);
};

// Helper function to get all category names
export const getCategoryNames = (): string[] => {
  return PRODUCT_CATEGORIES.map(category => category.displayName);
};

// Helper function to search products across all categories
export const searchProducts = (query: string): { category: ProductCategory; product: CatalogProduct }[] => {
  const results: { category: ProductCategory; product: CatalogProduct }[] = [];
  
  PRODUCT_CATEGORIES.forEach(category => {
    category.products.forEach(product => {
      if (
        product.description.toLowerCase().includes(query.toLowerCase()) ||
        product.modelNo.toLowerCase().includes(query.toLowerCase())
      ) {
        results.push({ category, product });
      }
    });
  });
  
  return results;
};
"""
    
    return ts_content

def main():
    """Main function"""
    file_path = r"c:\Users\Admin\OneDrive\Apps\baohui-baby-website\DATA\Phoenix Catlog7.xlsx"
    output_path = r"c:\Users\Admin\OneDrive\Apps\baohui-baby-website\src\data\catalogData.ts"
    
    if not os.path.exists(file_path):
        print(f"Error: Excel file not found: {file_path}")
        return False
    
    # Extract all products
    categories = extract_all_products(file_path)
    
    if not categories:
        print("No categories extracted. Exiting.")
        return False
    
    # Generate TypeScript file
    print("\n=== GENERATING TYPESCRIPT CATALOG ===")
    ts_content = generate_typescript_catalog(categories)
    
    # Write to file
    try:
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(ts_content)
        print(f"✓ Catalog updated successfully: {output_path}")
        
        # Print summary
        total_products = sum(len(cat['products']) for cat in categories)
        print(f"\n=== SUMMARY ===")
        print(f"Total categories: {len(categories)}")
        print(f"Total products: {total_products}")
        
        for category in categories:
            print(f"  {category['displayName']}: {len(category['products'])} products")
        
        return True
        
    except Exception as e:
        print(f"Error writing TypeScript file: {e}")
        return False

if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)