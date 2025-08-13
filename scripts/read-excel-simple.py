#!/usr/bin/env python3
"""
Simple script to read Phoenix Catalog 7 Excel file using openpyxl directly
"""

import sys
import os
from openpyxl import load_workbook

def read_excel_file(file_path):
    """Read the Excel file using openpyxl"""
    try:
        # Load the workbook
        workbook = load_workbook(file_path, data_only=True)
        
        print("=== PHOENIX CATALOG 7 ANALYSIS ===")
        print(f"File: {file_path}")
        print(f"Number of sheets: {len(workbook.sheetnames)}")
        print(f"Sheet names: {workbook.sheetnames}")
        print()
        
        # Analyze each sheet
        for sheet_name in workbook.sheetnames:
            print(f"=== SHEET: {sheet_name} ===")
            try:
                sheet = workbook[sheet_name]
                
                # Get dimensions
                max_row = sheet.max_row
                max_col = sheet.max_column
                print(f"Dimensions: {max_row} rows x {max_col} columns")
                
                # Get headers (first row)
                headers = []
                for col in range(1, min(max_col + 1, 20)):  # Limit to first 20 columns
                    cell_value = sheet.cell(row=1, column=col).value
                    headers.append(str(cell_value) if cell_value is not None else f"Col{col}")
                
                print(f"Headers: {headers}")
                
                # Show first few data rows
                print("\nFirst 5 data rows:")
                for row in range(2, min(7, max_row + 1)):  # Skip header, show next 5 rows
                    row_data = []
                    for col in range(1, min(max_col + 1, 10)):  # Limit to first 10 columns
                        cell_value = sheet.cell(row=row, column=col).value
                        row_data.append(str(cell_value) if cell_value is not None else "")
                    print(f"Row {row}: {row_data}")
                
                # Look for category information
                category_found = False
                for col in range(1, min(max_col + 1, 20)):
                    header = sheet.cell(row=1, column=col).value
                    if header and 'category' in str(header).lower():
                        print(f"\nCategory column found: {header}")
                        # Show unique categories
                        categories = set()
                        for row in range(2, min(max_row + 1, 100)):  # Check first 100 rows
                            cat_value = sheet.cell(row=row, column=col).value
                            if cat_value:
                                categories.add(str(cat_value))
                        print(f"Categories found: {sorted(categories)}")
                        category_found = True
                        break
                
                if not category_found:
                    print("No category column found in headers")
                
                print("\n" + "="*50 + "\n")
                
            except Exception as e:
                print(f"Error reading sheet {sheet_name}: {e}")
                print()
        
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return False
    
    return True

if __name__ == "__main__":
    file_path = r"c:\Users\Admin\OneDrive\Apps\baohui-baby-website\DATA\Phoenix Catlog7.xlsx"
    
    if not os.path.exists(file_path):
        print(f"File not found: {file_path}")
        sys.exit(1)
    
    read_excel_file(file_path)